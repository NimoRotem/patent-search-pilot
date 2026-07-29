"""Render the export model to a workbook (openpyxl).

WHY A SPREADSHEET AT ALL
------------------------
The PDF and the DOCX are a *document*: a cover, a narrative, a claim chart, prose per reference.
That is the right shape for something that gets filed or sent to counsel, and the wrong shape for
the thing people actually do with a result list first — sort it, filter it, hand it to someone
else, paste a column into an email. This export is the RESULT LIST, one row per reference, with
the same fields the card shows on screen and the card's own sketch embedded in the row, so the
sheet reads like the page it came from rather than like a database dump.

WHAT IT IS NOT
--------------
It is not a smaller PDF. It carries the same scope disclosure and the same verification-state
colouring as every other surface (`disclosure.py` is the single source for both) — a spreadsheet
looks authoritative in a way prose does not, so shipping one without the caveats would make the
weakest surface the most confident-looking one. The full claims and description are deliberately
NOT here: they belong in the Markdown export, which exists for exactly that.
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

import disclosure

# One visual system, matching the documents: navy headers, muted meta, semantic tones only where
# they mean something.
NAVY = "0B2545"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(color=NAVY, bold=True, size=14)
WARN_FONT = Font(color="B25E00", bold=True, size=10)
MUTED_FONT = Font(color="5B6B82", size=9)
BODY = Font(size=10)
THIN = Side(style="thin", color="D6DCE5")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Row geometry for the embedded sketch. 132x162 px is the card thumbnail; Excel row height is in
# points (1 px ~= 0.75 pt) and column width is in "characters" (~7 px each).
THUMB_PX = 118
ROW_PT = 116
THUMB_COL_W = 18

BASIS_LABEL = {"public_prior_art": "PUBLIC prior art",
               "secret_prior_art": "SECRET prior art (novelty only)",
               "priority_interval": "priority-interval art"}

# (header, width, extractor). One place defines the sheet, so adding a card field is one line.
COLUMNS = [
    ("Sketch",        THUMB_COL_W, lambda r: ""),                       # filled by the image pass
    ("#",              5, lambda r: r.get("rank")),
    ("Relevancy",     10, lambda r: r.get("relevancy")),
    ("Juris.",         7, lambda r: _juris(r)),
    ("Title",         42, lambda r: r.get("title") or "(untitled)"),
    ("Publication",   19, lambda r: r.get("pub")),
    ("Legal status",  15, lambda r: r.get("status_label") or ""),
    ("Prior-art basis", 17, lambda r: BASIS_LABEL.get(r.get("basis"), "not dated")),
    ("Priority",      11, lambda r: r.get("priority_date") or ""),
    ("Filed",         11, lambda r: r.get("filing_date") or ""),
    ("Published",     11, lambda r: r.get("publication_date") or ""),
    ("Assignee",      26, lambda r: "; ".join(r.get("assignees") or [])),
    ("Inventors",     26, lambda r: ", ".join(r.get("inventors") or [])),
    ("CPC",           20, lambda r: ", ".join(r.get("cpc") or [])),
    ("Family",        30, lambda r: _family_cell(r)),
    ("Found via",     20, lambda r: ", ".join(r.get("found_via") or [])),
    ("Figures",        8, lambda r: r.get("n_images") or 0),
    ("Reads on",      30, lambda r: " · ".join(r.get("covers_elements") or [])),
    ("Abstract",      60, lambda r: (r.get("abstract") or "")[:1800]),
    ("Why relevant",  60, lambda r: _why_cell(r)),
    ("Matched passage", 60, lambda r: _quoted_cell(r)),
    ("Google Patents", 34, lambda r: r.get("google_patents") or ""),
    ("Espacenet",     34, lambda r: r.get("espacenet") or ""),
    ("PDF",           34, lambda r: r.get("pdf_url") or ""),
]
LINK_HEADERS = {"Google Patents", "Espacenet", "PDF"}


def _juris(r):
    """A two-letter office code, always.

    `country` is the local corpus column for a corpus row but arrives from the source for a
    federated-only hit, where it can be a full country name ("United States"). One column cannot
    be half codes and half prose and still sort or filter, so prefer the publication number's own
    office prefix, which is the authoritative two letters by construction.
    """
    pub = (r.get("pub") or "").strip().upper()
    if len(pub) >= 2 and pub[:2].isalpha():
        return pub[:2]
    c = (r.get("country") or "").strip()
    return c.upper() if len(c) == 2 else c


def _family_cell(r):
    """"Family of 4 in 2 jurisdictions" + the year → jurisdiction strip, as one cell."""
    head = r.get("family_summary") or ""
    strip = "; ".join(f"{yr}: {', '.join(ccs)}" for yr, ccs in (r.get("family_timeline") or []) if yr)
    return "\n".join(x for x in (head, strip) if x)


def _why_cell(r):
    """The card's written opinion, then the grounded rationale — both, when they differ."""
    parts = []
    op = (r.get("relevancy_opinion") or "").strip()
    why = (r.get("why") or "").strip()
    if op:
        parts.append(op)
    if why and why != op:
        parts.append(why)
    if r.get("reads_on"):
        parts.append("Reads on: " + ", ".join(r["reads_on"]))
    return "\n\n".join(parts)


def _quoted_cell(r):
    q = r.get("quoted")
    if not q:
        return ""
    return f"[{q.get('kind')} {q.get('coord')} · similarity {q.get('score')}]\n{q.get('text') or ''}"


def _thumb_bytes(path, px=THUMB_PX):
    """Downscale a figure to a row-sized PNG in memory.

    Patent facsimiles are frequently 2000px+ TIFF/PNG sheets. Embedding them at native size makes a
    25-reference workbook tens of megabytes and Excel slow to open, so every sketch is resized once
    here. Returns None for anything Pillow cannot read — a missing thumbnail must never fail an
    export.
    """
    try:
        with PILImage.open(path) as im:
            im = im.convert("RGB")
            im.thumbnail((px, px), PILImage.LANCZOS)
            buf = BytesIO()
            im.save(buf, format="PNG")
            buf.seek(0)
            return buf, im.size
    except Exception:
        return None, None


def _header_row(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill, c.font, c.alignment, c.border = HEAD_FILL, HEAD_FONT, CENTER, BOX


def _results_sheet(wb, model):
    ws = wb.create_sheet("Results")
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"{disclosure.DOC_TITLE} — ranked references"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = disclosure.DOC_SUBTITLE
    ws["A2"].font = WARN_FONT
    ws["A3"] = f"Query: {model['query'][:2000]}"
    ws["A3"].font = MUTED_FONT
    ws["A4"] = (f"{model['mode'].replace('_', ' ').title()} mode · generated {model['generated']} · "
                f"{len(model['references'])} references exported of {model.get('n_cards') or 0} ranked")
    ws["A4"].font = MUTED_FONT

    head_row = 6
    _header_row(ws, head_row, [h for h, _w, _f in COLUMNS])
    for j, (_h, w, _f) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, r in enumerate(model["references"]):
        row = head_row + 1 + i
        for j, (h, _w, fn) in enumerate(COLUMNS, start=1):
            try:
                val = fn(r)
            except Exception:
                val = ""
            c = ws.cell(row=row, column=j, value=val)
            c.font, c.alignment, c.border = BODY, TOP_WRAP, BOX
            if h in LINK_HEADERS and val:
                c.hyperlink = val
                c.font = Font(size=9, color="0050A0", underline="single")
            elif h == "Relevancy":
                c.alignment = CENTER
                c.font = Font(size=11, bold=True)
            elif h in ("#", "Figures", "Juris."):
                c.alignment = CENTER
        ws.row_dimensions[row].height = ROW_PT

        #  ONE sketch per reference, in the row, exactly as the ask reads. `drawing_path` is
        #  resolved by export_data._best_drawing, which now also pulls down the lemad-Mongo CDN
        #  figure that most cards actually display — before that fix this column was empty for
        #  the majority of results while the screen showed a drawing for all of them.
        if r.get("drawing_path"):
            buf, size = _thumb_bytes(r["drawing_path"])
            if buf:
                img = XLImage(buf)
                img.width, img.height = size
                ws.add_image(img, f"A{row}")
            else:
                ws.cell(row=row, column=1, value="[figure unreadable]").font = MUTED_FONT
        else:
            ws.cell(row=row, column=1, value="[facsimile not digitized]").font = MUTED_FONT

    ws.freeze_panes = ws.cell(row=head_row + 1, column=6)
    ws.auto_filter.ref = (f"B{head_row}:{get_column_letter(len(COLUMNS))}"
                          f"{head_row + len(model['references'])}")
    return ws


def _chart_sheet(wb, model):
    """The element × reference retrieval map, coloured by VERIFICATION STATE (not by score)."""
    ws = wb.create_sheet("Element x reference")
    ws.sheet_view.showGridLines = False
    chart = model.get("claim_chart") or {}
    cols = chart.get("columns") or []

    ws["A1"] = disclosure.CHART_TITLE
    ws["A1"].font = TITLE_FONT
    ws["A2"] = disclosure.CHART_TAG.upper()
    ws["A2"].font = WARN_FONT
    ws["A3"] = disclosure.chart_warning()
    ws["A3"].font = MUTED_FONT
    ws["A3"].alignment = TOP_WRAP
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(2, len(cols) + 1))
    ws.row_dimensions[3].height = 58

    head = 5
    _header_row(ws, head, ["Invention element"] + [c.get("pub") for c in cols])
    ws.column_dimensions["A"].width = 46
    for j in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 20

    for i, row in enumerate(chart.get("rows") or []):
        r = head + 1 + i
        c0 = ws.cell(row=r, column=1, value=row.get("element"))
        c0.font, c0.alignment, c0.border = BODY, TOP_WRAP, BOX
        for j, cell in enumerate(row.get("cells") or [], start=2):
            if not cell.get("covered"):
                c = ws.cell(row=r, column=j, value="·")
                c.font, c.alignment, c.border = MUTED_FONT, CENTER, BOX
                continue
            _mark, _label, fill, _cov = disclosure.cell_state(cell)
            word = disclosure.cell_word(cell)
            bits = [word, str(cell.get("score"))]
            if cell.get("coord"):
                bits.append(str(cell["coord"]))
            c = ws.cell(row=r, column=j, value="\n".join(bits))
            c.fill = PatternFill("solid", fgColor=fill)
            c.font, c.alignment, c.border = Font(size=9, color="1B2A3A"), CENTER, BOX
        ws.row_dimensions[r].height = 40

    r = head + len(chart.get("rows") or []) + 2
    ws.cell(row=r, column=1, value="Legend").font = Font(bold=True, size=10)
    for word, n, gloss in disclosure.legend_lines(chart):
        r += 1
        ws.cell(row=r, column=1, value=f"{n} {word}").font = BODY
        ws.cell(row=r, column=2, value=gloss).font = MUTED_FONT
    summ = disclosure.verification_summary(chart)
    if summ:
        r += 2
        ws.cell(row=r, column=1, value=summ).font = Font(bold=True, size=10)
    ws.freeze_panes = ws.cell(row=head + 1, column=2)
    return ws


def _scope_sheet(wb, model):
    """Everything the three collapsed panels on the report page hold: what was searched, where,
    how far the search got, and how reliable that is."""
    ws = wb.create_sheet("Search scope")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    r = 1

    def kv(k, v, font=BODY):
        nonlocal r
        a = ws.cell(row=r, column=1, value=k); a.font = Font(bold=True, size=10); a.alignment = TOP_WRAP
        b = ws.cell(row=r, column=2, value=v); b.font = font; b.alignment = TOP_WRAP
        r += 1

    def head(t):
        nonlocal r
        r += 1
        c = ws.cell(row=r, column=1, value=t); c.font = TITLE_FONT
        r += 1

    ws.cell(row=r, column=1, value=disclosure.DOC_TITLE).font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value=disclosure.DOC_SUBTITLE).font = WARN_FONT
    r += 2

    head("What was searched")
    kv("Query (in full)", model["query"])
    kv("Mode", model["mode"].replace("_", " ").title())
    kv("Subject", str(model.get("subject") or "—"))
    kv("Generated", model["generated"])
    kv("Languages", ", ".join(model.get("languages") or []) or "—")
    dom = model.get("domain") or {}
    if dom.get("reason"):
        kv("Domain detector", dom["reason"], MUTED_FONT)

    head("Coverage ledger")
    kv("Rounds", model.get("rounds"))
    kv("Families surfaced", model.get("n_families_surfaced"))
    kv("References ranked", model.get("n_cards"))
    kv("Elements", f"{model['n_covered']}/{model['n_elements']} disclosed by the cited art")
    kv("Apparently novel", ", ".join(model.get("uncovered_elements") or []) or "none identified")
    kv("CPC branches searched", ", ".join(model.get("cpc_branches") or []) or "—")
    kv("Retrieval channels", ", ".join(model.get("channels_used") or []) or "—")
    kv("New families per round",
       ", ".join(f"r{i + 1}: {n}" for i, n in enumerate(model.get("round_new_families") or [])) or "—")

    if model.get("source_tags"):
        head("Sources searched")
        for s in model["source_tags"]:
            state = {"used": "used", "none": "no results", "failed": "failed",
                     "unknown": "not run", "off": "not configured"}.get(s.get("state"), s.get("state"))
            n = f" — {s['n']} hits" if s.get("n") else ""
            kv(s.get("label"), f"{state}{n}{('; ' + s['why']) if s.get('why') else ''}",
               BODY if s.get("state") == "used" else MUTED_FONT)

    head("Scope and measured reliability — read before relying on this workbook")
    for heading, body in disclosure.scope_paragraphs():
        kv(heading, body, Font(size=9.5))
    kv("Indexed CPC classes", "; ".join(disclosure.cpc_lines()), Font(size=9))
    kv("Not indexed", disclosure.not_indexed(), MUTED_FONT)
    kv("Method & corpus", model["corpus_note"], MUTED_FONT)
    return ws


def _appendix_sheet(wb, model):
    ws = wb.create_sheet("Full ranked list")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Every reference the search ranked, including those not selected for export"
    ws["A1"].font = TITLE_FONT
    heads = ["#", "Publication", "Title", "Match", "Basis", "Channels"]
    _header_row(ws, 3, heads)
    for w, col in zip([6, 20, 70, 10, 26, 30], "ABCDEF"):
        ws.column_dimensions[col].width = w
    for i, a in enumerate(model.get("appendix") or []):
        r = 4 + i
        vals = [a.get("rank"), a.get("pub"), a.get("title"), a.get("score"),
                BASIS_LABEL.get(a.get("basis"), "not dated"), ", ".join(a.get("channels") or [])]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.alignment, c.border = BODY, TOP_WRAP, BOX
    ws.freeze_panes = "A4"
    return ws


def render(model, out_path):
    wb = Workbook()
    wb.remove(wb.active)                       # drop the default empty sheet
    _results_sheet(wb, model)
    _chart_sheet(wb, model)
    _appendix_sheet(wb, model)
    _scope_sheet(wb, model)
    wb.active = 0
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return str(out_path)
